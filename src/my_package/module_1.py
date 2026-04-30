import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os

class InvoiceGenerator:
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb = openpyxl.load_workbook(template_path)
        self.ws = self.wb.active
        
    def generate_invoice(self, invoice_data):
        """
        invoice_data = {
            'invoice_no': 'AC/002/2026',
            'date': datetime.now(),
            'attn': 'David Chernitzky',
            'company': 'Armour Cyber LTD',
            'email': 'david.chernitzky@armourcyber.io; accounting@armourcyber.io',
            'items': [
                {'description': 'Delivery of services - December2025', 
                 'unit_price': 7129.3, 
                 'qty': 1}
            ],
            'currency': 'USD'
        }
        """
        
        # Fill header information
        self.ws['B3'] = invoice_data.get('attn', '')
        self.ws['B5'] = invoice_data.get('company', '')
        self.ws['E5'] = invoice_data.get('invoice_no', '')
        
        # Set date (cell I9 or similar — adjust based on your file)
        date_cell = self.ws['I9'] if self.ws['I9'].value == '=TODAY()' else self.ws['E9']
        date_cell.value = invoice_data.get('date', datetime.now()).strftime('%Y-%m-%d')
        
        # Email
        self.ws['E8'] = invoice_data.get('email', '')
        
        # Fill line items (starting row 16)
        row = 16
        for item in invoice_data.get('items', []):
            self.ws[f'B{row}'] = item['description']
            self.ws[f'D{row}'] = item['unit_price']
            self.ws[f'E{row}'] = item['qty']
            self.ws[f'F{row}'] = item['unit_price'] * item['qty']
            row += 1
        
        # Update totals
        sub_total_row = row + 2
        grand_total_row = sub_total_row + 1
        
        # Find the Sub Total and Grand Total rows by scanning
        for r in range(1, self.ws.max_row + 1):
            if self.ws[f'D{r}'].value == 'Sub Total':
                sub_total_row = r
            if self.ws[f'D{r}'].value == 'Grand Total (Nett) :':
                grand_total_row = r
        
        # Calculate totals
        total = 0
        for r in range(16, row):
            if self.ws[f'F{r}'].value and isinstance(self.ws[f'F{r}'].value, (int, float)):
                total += self.ws[f'F{r}'].value
        
        self.ws[f'F{sub_total_row}'] = total
        self.ws[f'F{grand_total_row}'] = total
        
        return self.wb
    
    def save_excel(self, output_path):
        self.wb.save(output_path)
        print(f"✅ Excel invoice saved to: {output_path}")
    
    def save_pdf(self, excel_path, pdf_path):
        """Convert Excel to PDF using optional library (requires additional setup)"""
        try:
            import win32com.client  # Windows only
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(excel_path))
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            wb.Close()
            excel.Quit()
            print(f"✅ PDF invoice saved to: {pdf_path}")
        except ImportError:
            print("⚠️ PDF generation requires pywin32 on Windows. Install: pip install pywin32")
        except Exception as e:
            print(f"⚠️ PDF generation failed: {e}")


# ======================
# USAGE EXAMPLE
# ======================
if __name__ == "__main__":
    # Sample invoice data
    invoice_data = {
        'invoice_no': 'AC/002/2026',
        'date': datetime.now(),
        'attn': 'David Chernitzky',
        'company': 'Armour Cyber LTD',
        'email': 'david.chernitzky@armourcyber.io; accounting@armourcyber.io',
        'items': [
            {
                'description': 'Delivery of services - December2025',
                'unit_price': 7129.30,
                'qty': 1
            }
        ]
    }
    
    # Initialize generator with your template
    generator = InvoiceGenerator('Invoice_Armour Cybersecurity_AC0022026.xlsx')
    
    # Generate filled invoice
    generator.generate_invoice(invoice_data)
    
    # Save as Excel
    generator.save_excel('Invoice_Generated_AC0022026.xlsx')
    
    # Optional: Save as PDF (Windows only, requires pywin32)
    # generator.save_pdf('Invoice_Generated_AC0022026.xlsx', 'Invoice_Generated_AC0022026.pdf')
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from fpdf import FPDF

class InvoiceGeneratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Professional Invoice Generator")
        self.root.geometry("1200x700")
        self.root.configure(bg='#f0f0f0')
        
        # Store items dynamically
        self.items = []
        self.item_rows = []
        
        # Setup GUI
        self.setup_gui()
        
    def setup_gui(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # ==================== HEADER SECTION ====================
        header_frame = ttk.LabelFrame(main_frame, text="Invoice Header Information", padding="10")
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,10))
        
        # Left side - From (Company Info)
        from_frame = ttk.Frame(header_frame)
        from_frame.grid(row=0, column=0, padx=(0,20))
        
        ttk.Label(from_frame, text="FROM (Your Company):", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0,5))
        
        self.from_name = tk.Text(from_frame, height=2, width=35)
        self.from_name.grid(row=1, column=0, pady=2)
        self.from_name.insert('1.0', "Mohd Fairuz Bin Ibrahim\nShopIQ S/B")
        
        self.from_address = tk.Text(from_frame, height=3, width=35)
        self.from_address.grid(row=2, column=0, pady=2)
        self.from_address.insert('1.0', "21st Floor, CMA Building\n64 Connaught Road Central\nHong Kong")
        
        self.from_email = ttk.Entry(from_frame, width=35)
        self.from_email.grid(row=3, column=0, pady=2)
        self.from_email.insert(0, "fairuz@gospurr.com")
        self.from_email.insert(0, "fairuz@gospurr.com")
        
        # Right side - To (Customer Info)
        to_frame = ttk.Frame(header_frame)
        to_frame.grid(row=0, column=1, sticky=(tk.W, tk.E))
        
        ttk.Label(to_frame, text="TO (Customer):", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0,5))
        
        ttk.Label(to_frame, text="Attention:").grid(row=1, column=0, sticky=tk.W)
        self.attn = ttk.Entry(to_frame, width=35)
        self.attn.grid(row=2, column=0, pady=2)
        
        ttk.Label(to_frame, text="Company:").grid(row=3, column=0, sticky=tk.W)
        self.company = ttk.Entry(to_frame, width=35)
        self.company.grid(row=4, column=0, pady=2)
        
        ttk.Label(to_frame, text="Email:").grid(row=5, column=0, sticky=tk.W)
        self.customer_email = ttk.Entry(to_frame, width=35)
        self.customer_email.grid(row=6, column=0, pady=2)
        
        # Invoice Details
        invoice_frame = ttk.LabelFrame(main_frame, text="Invoice Details", padding="10")
        invoice_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0,10))
        
        ttk.Label(invoice_frame, text="Invoice No:").grid(row=0, column=0, padx=5)
        self.invoice_no = ttk.Entry(invoice_frame, width=20)
        self.invoice_no.grid(row=0, column=1, padx=5)
        self.invoice_no.insert(0, f"AC/{datetime.now().strftime('%m')}/{datetime.now().year}")
        
        ttk.Label(invoice_frame, text="Date:").grid(row=0, column=2, padx=5)
        self.invoice_date = ttk.Entry(invoice_frame, width=15)
        self.invoice_date.grid(row=0, column=3, padx=5)
        self.invoice_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        ttk.Label(invoice_frame, text="Currency:").grid(row=0, column=4, padx=5)
        self.currency = ttk.Combobox(invoice_frame, values=["USD", "EUR", "GBP", "HKD", "SGD"], width=10)
        self.currency.grid(row=0, column=5, padx=5)
        self.currency.set("USD")
        
        # ==================== ITEMS SECTION (DYNAMIC) ====================
        items_frame = ttk.LabelFrame(main_frame, text="Invoice Items", padding="10")
        items_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0,10))
        
        # Scrollable area for items
        self.items_canvas = tk.Canvas(items_frame, height=250)
        scrollbar = ttk.Scrollbar(items_frame, orient="vertical", command=self.items_canvas.yview)
        self.items_scrollable = ttk.Frame(self.items_canvas)
        
        self.items_scrollable.bind("<Configure>", lambda e: self.items_canvas.configure(scrollregion=self.items_canvas.bbox("all")))
        self.items_canvas.create_window((0, 0), window=self.items_scrollable, anchor="nw")
        self.items_canvas.configure(yscrollcommand=scrollbar.set)
        
        self.items_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Header for items table
        ttk.Label(self.items_scrollable, text="Description", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(self.items_scrollable, text="Unit Price", font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(self.items_scrollable, text="Quantity", font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5, pady=5)
        ttk.Label(self.items_scrollable, text="Total", font=('Arial', 10, 'bold')).grid(row=0, column=3, padx=5, pady=5)
        ttk.Label(self.items_scrollable, text="", font=('Arial', 10, 'bold')).grid(row=0, column=4, padx=5, pady=5)
        
        # Button to add new item
        self.add_item_button = ttk.Button(items_frame, text="+ Add New Item", command=self.add_item_row)
        self.add_item_button.pack(pady=5)
        
        # Add initial 3 empty rows
        for _ in range(3):
            self.add_item_row()
        
        # ==================== TOTALS SECTION ====================
        totals_frame = ttk.LabelFrame(main_frame, text="Summary", padding="10")
        totals_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0,10))
        
        ttk.Label(totals_frame, text="Subtotal:", font=('Arial', 10)).grid(row=0, column=0, padx=5, pady=2, sticky=tk.E)
        self.subtotal_label = ttk.Label(totals_frame, text="0.00", font=('Arial', 10, 'bold'), foreground='blue')
        self.subtotal_label.grid(row=0, column=1, padx=5, pady=2, sticky=tk.W)
        
        ttk.Label(totals_frame, text="Total (Nett):", font=('Arial', 12, 'bold')).grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
        self.grand_total_label = ttk.Label(totals_frame, text="0.00", font=('Arial', 12, 'bold'), foreground='green')
        self.grand_total_label.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        
        # ==================== BANK DETAILS SECTION ====================
        bank_frame = ttk.LabelFrame(main_frame, text="Bank Transfer Details", padding="10")
        bank_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,10))
        
        self.bank_details = tk.Text(bank_frame, height=6, width=80)
        self.bank_details.pack()
        self.bank_details.insert('1.0', "Account Name: Mohd Fairuz Bin Ibrahim\nAccount No.: 121-255707-833\nBIC/SWIFT code: HSBCHKHHHKH\nBank Name: HSBC Bank, Tsim Sha Tsui Premier Centre\nAddress: 2/F & 3/F, HSBC Building Tsim Sha Tsui, 82-84 Nathan Road, Tsim Sha Tsui, Kowloon Hong Kong")
        
        # ==================== BUTTONS ====================
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Generate Excel", command=self.generate_excel, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Generate PDF", command=self.generate_pdf, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear All", command=self.clear_all, width=15).pack(side=tk.LEFT, padx=5)
        
    def add_item_row(self):
        """Dynamically add a new item row"""
        row_num = len(self.item_rows)
        frame = ttk.Frame(self.items_scrollable)
        frame.grid(row=row_num + 1, column=0, columnspan=5, sticky=(tk.W, tk.E), pady=2)
        
        desc_entry = ttk.Entry(frame, width=50)
        desc_entry.grid(row=0, column=0, padx=5)
        
        price_entry = ttk.Entry(frame, width=15)
        price_entry.grid(row=0, column=1, padx=5)
        price_entry.bind('<KeyRelease>', lambda e, r=row_num: self.update_item_total(r))
        
        qty_entry = ttk.Entry(frame, width=10)
        qty_entry.grid(row=0, column=2, padx=5)
        qty_entry.bind('<KeyRelease>', lambda e, r=row_num: self.update_item_total(r))
        
        total_label = ttk.Label(frame, text="0.00", width=15, foreground='blue')
        total_label.grid(row=0, column=3, padx=5)
        
        remove_btn = ttk.Button(frame, text="✖ Remove", command=lambda f=frame, r=row_num: self.remove_item_row(r, f))
        remove_btn.grid(row=0, column=4, padx=5)
        
        self.item_rows.append({
            'frame': frame,
            'desc': desc_entry,
            'price': price_entry,
            'qty': qty_entry,
            'total': total_label
        })
        
    def remove_item_row(self, row_index, frame):
        """Remove an item row"""
        if len(self.item_rows) > 1:  # Keep at least one row
            frame.destroy()
            self.item_rows.pop(row_index)
            self.update_all_totals()
            
    def update_item_total(self, row_index):
        """Calculate total for a single item"""
        try:
            price = float(self.item_rows[row_index]['price'].get() or 0)
            qty = float(self.item_rows[row_index]['qty'].get() or 0)
            total = price * qty
            self.item_rows[row_index]['total'].config(text=f"{total:,.2f}")
            self.update_all_totals()
        except ValueError:
            self.item_rows[row_index]['total'].config(text="0.00")
            
    def update_all_totals(self):
        """Calculate subtotal and grand total"""
        subtotal = 0
        for item in self.item_rows:
            try:
                price = float(item['price'].get() or 0)
                qty = float(item['qty'].get() or 0)
                subtotal += price * qty
            except ValueError:
                pass
                
        self.subtotal_label.config(text=f"{subtotal:,.2f}")
        self.grand_total_label.config(text=f"{subtotal:,.2f}")
        
    def get_items_list(self):
        """Return list of items for export"""
        items = []
        for item in self.item_rows:
            desc = item['desc'].get().strip()
            if desc:  # Only include non-empty descriptions
                try:
                    price = float(item['price'].get() or 0)
                    qty = float(item['qty'].get() or 0)
                    if price > 0 and qty > 0:
                        items.append({
                            'description': desc,
                            'unit_price': price,
                            'qty': qty,
                            'total': price * qty
                        })
                except ValueError:
                    pass
        return items
    
    def generate_excel(self):
        """Generate Excel invoice"""
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoice"
            
            # Style definitions
            header_font = Font(bold=True, size=12)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Company info
            ws['A1'] = self.from_name.get('1.0', 'end-1c')
            ws['A2'] = self.from_address.get('1.0', 'end-1c')
            ws['A3'] = f"Email: {self.from_email.get()}"
            
            # Customer info
            ws['E1'] = "INVOICE"
            ws['E1'].font = Font(bold=True, size=16)
            ws['E3'] = f"Attn: {self.attn.get()}"
            ws['E4'] = f"Company: {self.company.get()}"
            ws['E5'] = f"Invoice No: {self.invoice_no.get()}"
            ws['E6'] = f"Date: {self.invoice_date.get()}"
            
            # Items table header
            row = 10
            headers = ['Description', 'Unit Price', 'Qty', 'Total']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.border = border
                
            # Items data
            items = self.get_items_list()
            for idx, item in enumerate(items, start=1):
                ws.cell(row=row+idx, column=1, value=item['description']).border = border
                ws.cell(row=row+idx, column=2, value=item['unit_price']).border = border
                ws.cell(row=row+idx, column=3, value=item['qty']).border = border
                ws.cell(row=row+idx, column=4, value=item['total']).border = border
                
            # Totals
            total_row = row + len(items) + 2
            ws.cell(row=total_row, column=3, value="Subtotal:").font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=float(self.subtotal_label['text'].replace(',', '')))
            
            grand_row = total_row + 1
            ws.cell(row=grand_row, column=3, value="Grand Total:").font = Font(bold=True, size=12)
            ws.cell(row=grand_row, column=4, value=float(self.grand_total_label['text'].replace(',', ''))).font = Font(bold=True)
            
            # Bank details
            bank_row = grand_row + 3
            ws.cell(row=bank_row, column=1, value="Bank Transfer Details:").font = Font(bold=True)
            bank_lines = self.bank_details.get('1.0', 'end-1c').split('\n')
            for i, line in enumerate(bank_lines):
                ws.cell(row=bank_row + i + 1, column=1, value=line)
                
            # Save file
            filename = f"Invoice_{self.invoice_no.get()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename)
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Success", f"Excel invoice saved to:\n{filepath}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel: {str(e)}")
            
    def generate_pdf(self):
        """Generate PDF invoice"""
        try:
            class PDF(FPDF):
                def header(self):
                    self.set_font('Arial', 'B', 16)
                    self.cell(0, 10, 'INVOICE', 0, 1, 'R')
                    self.ln(5)
                    
            pdf = PDF()
            pdf.add_page()
            pdf.set_font('Arial', '', 10)
            
            # From info
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 6, "FROM:", 0, 1)
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 5, self.from_name.get('1.0', 'end-1c'))
            pdf.multi_cell(0, 5, self.from_address.get('1.0', 'end-1c'))
            pdf.cell(0, 5, f"Email: {self.from_email.get()}", 0, 1)
            pdf.ln(5)
            
            # Customer info
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 6, "TO:", 0, 1)
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 5, f"Attn: {self.attn.get()}", 0, 1)
            pdf.cell(0, 5, f"Company: {self.company.get()}", 0, 1)
            pdf.cell(0, 5, f"Email: {self.customer_email.get()}", 0, 1)
            pdf.ln(5)
            
            # Invoice details
            pdf.cell(60, 5, f"Invoice No: {self.invoice_no.get()}", 0, 0)
            pdf.cell(60, 5, f"Date: {self.invoice_date.get()}", 0, 1)
            pdf.cell(60, 5, f"Currency: {self.currency.get()}", 0, 1)
            pdf.ln(5)
            
            # Items table
            pdf.set_fill_color(230, 230, 230)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(80, 10, 'Description', 1, 0, 'C', 1)
            pdf.cell(35, 10, 'Unit Price', 1, 0, 'C', 1)
            pdf.cell(30, 10, 'Qty', 1, 0, 'C', 1)
            pdf.cell(35, 10, 'Total', 1, 1, 'C', 1)
            
            pdf.set_font('Arial', '', 10)
            items = self.get_items_list()
            for item in items:
                pdf.cell(80, 8, item['description'][:40], 1)
                pdf.cell(35, 8, f"{self.currency.get()} {item['unit_price']:,.2f}", 1, 0, 'R')
                pdf.cell(30, 8, str(item['qty']), 1, 0, 'C')
                pdf.cell(35, 8, f"{self.currency.get()} {item['total']:,.2f}", 1, 1, 'R')
                
            # Grand total
            total = float(self.grand_total_label['text'].replace(',', ''))
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(145, 10, 'Grand Total (Nett):', 1, 0, 'R', 1)
            pdf.cell(35, 10, f"{self.currency.get()} {total:,.2f}", 1, 1, 'C', 1)
            
            # Bank details
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 6, "Bank Transfer Details:", 0, 1)
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 4, self.bank_details.get('1.0', 'end-1c'))
            
            # Signature
            pdf.ln(5)
            pdf.cell(0, 6, "On Behalf of ShopIQ S/B", 0, 1)
            pdf.ln(5)
            pdf.cell(0, 6, "Mohd Fairuz Ibrahim", 0, 1)
            pdf.cell(0, 6, "Director", 0, 1)
            
            # Save PDF
            filename = f"Invoice_{self.invoice_no.get()}_{datetime.now().strftime('%Y%m%d')}.pdf"
            filepath = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=filename)
            if filepath:
                pdf.output(filepath)
                messagebox.showinfo("Success", f"PDF invoice saved to:\n{filepath}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            
    def clear_all(self):
        """Clear all fields"""
        if messagebox.askyesno("Confirm", "Clear all invoice data?"):
            self.attn.delete(0, tk.END)
            self.company.delete(0, tk.END)
            self.customer_email.delete(0, tk.END)
            
            # Clear items
            for item in self.item_rows[:]:
                item['frame'].destroy()
            self.item_rows.clear()
            
            # Add new empty rows
            for _ in range(3):
                self.add_item_row()
                
            self.update_all_totals()
            messagebox.showinfo("Success", "All fields cleared!")


# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceGeneratorGUI(root)
    root.mainloop()